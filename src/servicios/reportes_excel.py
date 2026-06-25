from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from src.utils.decoradores import medir_tiempo, registrar_log


VENTAS_COLUMNAS = [
    # Columnas internas: mantienen nombres estables para el codigo.
    # Los encabezados visibles se traducen con VENTAS_ENCABEZADOS.
    "pedido_id",
    "fecha",
    "cliente",
    "tipo_entrega",
    "direccion",
    "productos",
    "cantidad_total_productos",
    "subtotal_bruto_pedido",
    "descuento_pedido",
    "total_vendido_pedido",
    "total_vendido_hasta_el_momento",
]

VENTAS_ENCABEZADOS = {
    "pedido_id": "ID",
    "fecha": "Fecha",
    "cliente": "Cliente",
    "tipo_entrega": "Entrega",
    "direccion": "Dirección",
    "productos": "Productos",
    "cantidad_total_productos": "Cant. total",
    "subtotal_bruto_pedido": "Subtotal",
    "descuento_pedido": "Descuento",
    "total_vendido_pedido": "Total pedido",
    "total_vendido_hasta_el_momento": "Total acumulado",
}

STOCK_COLUMNAS = [
    # Columnas internas del reporte de stock.
    # La interfaz y Excel muestran nombres mas claros mediante STOCK_ENCABEZADOS.
    "ingrediente",
    "cantidad",
    "precio_unitario",
    "valor_total_stock",
    "estado",
]

STOCK_ENCABEZADOS = {
    "ingrediente": "Ingrediente",
    "cantidad": "Cantidad",
    "precio_unitario": "Precio unitario",
    "valor_total_stock": "Valor total",
    "estado": "Estado",
}


def _encabezado_visible(encabezado):
    # Permite leer un Excel generado y devolver encabezados bonitos en la interfaz.
    encabezados = {}
    encabezados.update(VENTAS_ENCABEZADOS)
    encabezados.update(STOCK_ENCABEZADOS)
    return encabezados.get(encabezado, encabezado)


def _nombre_ingrediente_visible(nombre):
    # Corrige nombres tecnicos de ingredientes antes de mostrarlos en reportes.
    ingredientes = {
        "jamon": "jamón",
        "morron": "morrón",
        "jamon_queso": "jamón y queso",
        "tapas_empanada": "tapas de empanada",
    }
    texto = str(nombre).strip()
    texto = ingredientes.get(texto.lower(), texto.replace("_", " "))
    return texto[0].upper() + texto[1:] if texto else ""


def obtener_carpeta_reportes():
    # Ubicacion unica para reportes Excel. Se crea sola si el usuario borra la carpeta.
    ruta_proyecto = Path(__file__).resolve().parents[2]
    carpeta_reportes = ruta_proyecto / "reportes"
    carpeta_reportes.mkdir(parents=True, exist_ok=True)
    return carpeta_reportes


def obtener_ruta_reporte(nombre_archivo):
    # Arma la ruta final de un reporte manteniendo todos los Excel juntos.
    carpeta_reportes = obtener_carpeta_reportes()
    return carpeta_reportes / nombre_archivo


def _convertir_entero(valor):
    # Conversor defensivo: si un dato viejo o incompleto no sirve, devuelve None.
    try:
        return int(valor)
    except (TypeError, ValueError):
        return None


def _convertir_numero(valor):
    # Normaliza importes y cantidades antes de agregarlos al reporte.
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def _convertir_fecha(valor):
    # Acepta fechas reales de Python o textos ISO guardados en JSON.
    if isinstance(valor, datetime):
        return valor

    if not valor:
        return None

    try:
        return datetime.fromisoformat(str(valor))
    except ValueError:
        return None


def _formato_moneda(valor):
    # Texto de dinero usado dentro de celdas descriptivas del reporte de ventas.
    texto = f"{valor:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"${texto}"


def transformar_ventas(ventas):
    # Limpia y normaliza las ventas antes de agruparlas.
    # Tambien evita duplicados simples para que el reporte no infle totales.
    filas = []
    claves_vistas = set()

    for venta in ventas:
        pedido_id = _convertir_entero(venta.get("pedido_id"))
        fecha = _convertir_fecha(venta.get("fecha"))
        cliente = str(venta.get("cliente", "")).strip()
        tipo_entrega = str(venta.get("tipo_entrega", "Retiro")).strip() or "Retiro"
        direccion = str(venta.get("direccion", "")).strip()
        producto = str(venta.get("producto", "")).strip()
        cantidad = _convertir_entero(venta.get("cantidad"))
        precio_unitario = _convertir_numero(venta.get("precio_unitario"))
        subtotal = _convertir_numero(venta.get("subtotal"))
        descuento = _convertir_numero(venta.get("descuento", 0))
        subtotal_bruto = _convertir_numero(venta.get("subtotal_bruto"))

        if None in (pedido_id, fecha, cantidad, precio_unitario, subtotal):
            continue

        if descuento is None:
            descuento = 0

        if subtotal_bruto is None:
            subtotal_bruto = subtotal + descuento

        if not producto:
            continue

        clave = (pedido_id, producto, cantidad)
        if clave in claves_vistas:
            continue

        claves_vistas.add(clave)
        filas.append(
            {
                "pedido_id": pedido_id,
                "fecha": fecha,
                "cliente": cliente,
                "tipo_entrega": tipo_entrega,
                "direccion": direccion,
                "producto": producto,
                "cantidad": cantidad,
                "precio_unitario": precio_unitario,
                "subtotal_bruto": subtotal_bruto,
                "descuento": descuento,
                "subtotal": subtotal,
            }
        )

    filas.sort(key=lambda fila: (fila["fecha"], fila["pedido_id"]))
    return filas


def transformar_stock(stock):
    # Acepta stock en dos formatos: lista detallada o diccionario simple.
    # Devuelve siempre filas con valor total y estado visual.
    if isinstance(stock, list):
        filas_origen = stock
    else:
        filas_origen = []
        for ingrediente, cantidad in stock.items():
            filas_origen.append(
                {
                    "ingrediente": ingrediente,
                    "cantidad": cantidad,
                    "precio_unitario": 0,
                }
            )

    filas = []
    for fila_origen in filas_origen:
        ingrediente = str(fila_origen.get("ingrediente", "")).strip()
        cantidad = _convertir_numero(fila_origen.get("cantidad"))
        precio_unitario = _convertir_numero(fila_origen.get("precio_unitario"))

        if not ingrediente:
            continue

        if cantidad is None:
            cantidad = 0

        if precio_unitario is None:
            precio_unitario = 0

        valor_total = cantidad * precio_unitario
        estado = "Reponer" if cantidad <= 5 else "Disponible"
        filas.append(
            {
                "ingrediente": _nombre_ingrediente_visible(ingrediente),
                "cantidad": cantidad,
                "precio_unitario": precio_unitario,
                "valor_total_stock": valor_total,
                "estado": estado,
            }
        )

    filas.sort(key=lambda fila: fila["ingrediente"])
    return filas


def _agrupar_ventas_por_pedido(filas_ventas):
    # Junta varias lineas de venta bajo un mismo pedido.
    # Asi el Excel muestra una fila por pedido y no una fila suelta por producto.
    pedidos = {}

    for venta in filas_ventas:
        pedido_id = venta["pedido_id"]
        if pedido_id not in pedidos:
            pedidos[pedido_id] = {
                "pedido_id": pedido_id,
                "fecha": venta["fecha"],
                "cliente": venta["cliente"],
                "tipo_entrega": venta["tipo_entrega"],
                "direccion": venta["direccion"],
                "productos": {},
            }

        productos = pedidos[pedido_id]["productos"]
        nombre_producto = venta["producto"]
        if nombre_producto not in productos:
            productos[nombre_producto] = {"cantidad": 0, "subtotal_bruto": 0, "descuento": 0, "subtotal": 0}

        productos[nombre_producto]["cantidad"] += venta["cantidad"]
        productos[nombre_producto]["subtotal_bruto"] += venta["subtotal_bruto"]
        productos[nombre_producto]["descuento"] += venta["descuento"]
        productos[nombre_producto]["subtotal"] += venta["subtotal"]

    return sorted(pedidos.values(), key=lambda pedido: (pedido["fecha"], pedido["pedido_id"]))


def armar_reporte_ventas(ventas):
    # Construye el contenido final de la hoja Ventas:
    # productos resumidos, descuentos, total por pedido y total acumulado.
    filas_ventas = transformar_ventas(ventas)

    if not filas_ventas:
        return [{"mensaje": "Todavía no existen ventas registradas."}]

    filas_reporte = []
    acumulado = 0

    for pedido in _agrupar_ventas_por_pedido(filas_ventas):
        textos_productos = []
        cantidad_total = 0
        subtotal_bruto_pedido = 0
        descuento_pedido = 0
        total_pedido = 0

        for nombre_producto, datos in pedido["productos"].items():
            cantidad_total += datos["cantidad"]
            subtotal_bruto_pedido += datos["subtotal_bruto"]
            descuento_pedido += datos["descuento"]
            total_pedido += datos["subtotal"]
            texto_producto = f"{nombre_producto} x{datos['cantidad']} ({_formato_moneda(datos['subtotal'])})"
            if datos["descuento"]:
                texto_producto += f" desc. {_formato_moneda(datos['descuento'])}"
            textos_productos.append(texto_producto)

        acumulado += total_pedido
        filas_reporte.append(
            {
                "pedido_id": pedido["pedido_id"],
                "fecha": pedido["fecha"],
                "cliente": pedido["cliente"],
                "tipo_entrega": pedido["tipo_entrega"],
                "direccion": pedido["direccion"],
                "productos": ", ".join(textos_productos),
                "cantidad_total_productos": cantidad_total,
                "subtotal_bruto_pedido": subtotal_bruto_pedido,
                "descuento_pedido": descuento_pedido,
                "total_vendido_pedido": total_pedido,
                "total_vendido_hasta_el_momento": acumulado,
            }
        )

    return filas_reporte


def _crear_libro(nombre_hoja, columnas, filas, encabezados):
    # Crea un libro Excel desde filas ya preparadas y aplica nombres visibles.
    # Si no hay datos, cambia a una hoja con mensaje de estado.
    libro = Workbook()
    hoja = libro.active
    hoja.title = nombre_hoja

    columnas_internas = columnas
    encabezados_visibles = [encabezados.get(columna, columna) for columna in columnas]
    if filas and "mensaje" in filas[0]:
        columnas_internas = ["mensaje"]
        encabezados_visibles = ["Estado"]

    hoja.append(encabezados_visibles)
    for fila in filas:
        hoja.append([fila.get(columna, "") for columna in columnas_internas])

    _aplicar_estilos(hoja, nombre_hoja)
    return libro


def _aplicar_estilos(hoja, nombre_tabla):
    # Estiliza el Excel para que parezca un reporte terminado:
    # encabezado oscuro, filas ajustables, columnas con ancho automatico y tabla.
    encabezado_fill = PatternFill("solid", fgColor="111827")
    encabezado_font = Font(color="FFFFFF", bold=True)

    for celda in hoja[1]:
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal="center")

    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)

    for columna in hoja.columns:
        ancho = 12
        letra = columna[0].column_letter
        for celda in columna:
            valor = "" if celda.value is None else str(celda.value)
            ancho = max(ancho, min(len(valor) + 2, 42))
        hoja.column_dimensions[letra].width = ancho

    hoja.freeze_panes = "A2"

    ultima_fila = hoja.max_row
    ultima_columna = hoja.max_column
    referencia = f"A1:{hoja.cell(row=ultima_fila, column=ultima_columna).coordinate}"
    tabla = Table(displayName=f"Tabla{nombre_tabla}", ref=referencia)
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tabla.tableStyleInfo = estilo
    hoja.add_table(tabla)


@registrar_log
@medir_tiempo
def generar_reporte_ventas(ventas, nombre_archivo="reporte_ventas.xlsx"):
    # Entrada publica para exportar ventas. Los decoradores dejan registro y tiempo.
    filas_reporte = armar_reporte_ventas(ventas)
    destino = obtener_ruta_reporte(nombre_archivo)
    libro = _crear_libro("Ventas", VENTAS_COLUMNAS, filas_reporte, VENTAS_ENCABEZADOS)
    libro.save(destino)
    return str(destino)


@registrar_log
@medir_tiempo
def generar_reporte_stock(stock, nombre_archivo="reporte_stock.xlsx"):
    # Entrada publica para exportar stock con cantidades, precios y estado.
    filas_reporte = transformar_stock(stock)
    destino = obtener_ruta_reporte(nombre_archivo)
    libro = _crear_libro("Stock", STOCK_COLUMNAS, filas_reporte, STOCK_ENCABEZADOS)
    libro.save(destino)
    return str(destino)


def leer_reporte_excel(nombre_archivo):
    # Lee un Excel ya generado para mostrarlo dentro de Tkinter sin usar DataFrames.
    ruta_reporte = obtener_ruta_reporte(nombre_archivo)

    if not ruta_reporte.exists():
        raise FileNotFoundError(f"No existe el reporte '{nombre_archivo}'. Primero debe generarlo.")

    libro = load_workbook(ruta_reporte, data_only=True)
    hoja = libro.active
    filas = list(hoja.iter_rows(values_only=True))

    if not filas:
        return []

    encabezados = [_encabezado_visible(str(valor)) if valor is not None else "" for valor in filas[0]]
    resultado = []

    for valores in filas[1:]:
        if all(valor is None for valor in valores):
            continue

        fila = {}
        for indice, encabezado in enumerate(encabezados):
            if not encabezado:
                continue

            valor = valores[indice] if indice < len(valores) else ""
            fila[encabezado] = valor

        resultado.append(fila)
    return resultado
